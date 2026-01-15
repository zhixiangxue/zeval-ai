"""Generate test dataset from PDF and export to Excel

This script:
1. Read PDF file using DoclingReader
2. Split into units using MarkdownHeaderSplitter
3. Enrich units with extractors (summary, keyphrases, entities)
4. Generate personas
5. Generate single-hop Q&A pairs
6. Export to Excel (same format as reporter's Case Details sheet)

Prerequisites:
    - OPENAI_API_KEY environment variable must be set
    - openpyxl installed for Excel export

Usage:
    python examples/test_generate_dataset.py
    
    Follow the interactive prompts to:
    - Enter PDF file path
    - Specify number of test cases
    - Specify maximum units for generation
    
    Output Excel file will be saved in the same directory as the PDF
    with the same name (but .xlsx extension).
"""

import asyncio
import os
import sys
import ssl
import logging

from pathlib import Path
from datetime import datetime
from dotenv import load_dotenv
from rich.console import Console
from rich.panel import Panel
from pydantic import BaseModel, Field

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("[ERROR] openpyxl not installed. Please run: pip install openpyxl")
    sys.exit(1)

# Disable SSL verification for HuggingFace downloads
ssl._create_default_https_context = ssl._create_unverified_context

# Suppress HTTP request logs
logging.getLogger("httpx").setLevel(logging.WARNING)
logging.getLogger("httpcore").setLevel(logging.WARNING)
logging.getLogger("openai").setLevel(logging.WARNING)

# Load environment variables FIRST
load_dotenv()

# Import required components (after dotenv)
from zeval.synthetic_data.readers.docling import DoclingReader
from zeval.synthetic_data.splitters.markdown import MarkdownHeaderSplitter
from zeval.synthetic_data.transforms.extractors import (
    SummaryExtractor,
    KeyphrasesExtractor,
    EntitiesExtractor
)
from zeval.synthetic_data.generators.persona import generate_personas, Persona
from zeval.synthetic_data.generators.single_hop import generate_single_hop
from zeval.schemas.eval import EvalDataset

console = Console()


# ============================================================================
# Domain-Specific Persona (customize for your domain)
# ============================================================================

class HomeBuyerPersona(Persona):
    """US Home Buyer persona with financial attributes"""
    credit_score: int = Field(
        description="Credit score (300-850), affects mortgage eligibility and interest rates"
    )
    dti_ratio: float = Field(
        description="Debt-to-Income ratio as percentage (typical max is 43%)"
    )
    down_payment_percent: float = Field(
        description="Down payment as percentage of home price (typically 3-20%)"
    )
    budget_range: str = Field(
        description="Home price budget range (e.g., '$300K-$500K')"
    )


# ============================================================================
# Excel Export Function
# ============================================================================

def export_dataset_to_excel(dataset: EvalDataset, output_path: Path):
    """Export dataset to Excel with same format as reporter's Case Details sheet"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Dataset"
    
    # Define border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Header (same as reporter's Case Details sheet)
    headers = [
        'Case ID',
        'Persona Name',
        'Persona Role',
        'Question',
        'Ground Truth Answer',
        'Ground Truth Contexts',
        'Source Unit IDs',
        'Generation Params'
    ]
    ws.append(headers)
    
    # Style header
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="424242", end_color="424242", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    
    # Add data
    for i, case in enumerate(dataset.cases, 1):
        # Extract persona info
        if case.persona:
            persona_name = case.persona.get('name', 'N/A')
            persona_role = case.persona.get('role_description', 'N/A')
        else:
            persona_name = 'N/A'
            persona_role = 'N/A'
        
        # Format ground truth contexts as numbered list
        ground_truth_contexts_text = "\n".join([
            f"{idx+1}. {ctx}" for idx, ctx in enumerate(case.ground_truth_contexts)
        ])
        
        # Format source unit IDs
        source_unit_ids = ", ".join([unit.get('unit_id', 'N/A') for unit in case.source_units])
        
        # Format generation params
        generation_params = f"hop_type: {case.generation_params.get('hop_type', 'N/A')}"
        
        # Build row
        row = [
            i,
            persona_name,
            persona_role,
            case.question,
            case.ground_truth_answer,
            ground_truth_contexts_text,
            source_unit_ids,
            generation_params
        ]
        
        ws.append(row)
        
        # Add borders to all cells
        row_idx = ws.max_row
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            
            # Enable text wrapping for long text fields
            if col_idx >= 4:  # Question and onwards
                cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 10   # Case ID
    ws.column_dimensions['B'].width = 20   # Persona Name
    ws.column_dimensions['C'].width = 30   # Persona Role
    ws.column_dimensions['D'].width = 60   # Question
    ws.column_dimensions['E'].width = 60   # Ground Truth Answer
    ws.column_dimensions['F'].width = 80   # Ground Truth Contexts
    ws.column_dimensions['G'].width = 20   # Source Unit IDs
    ws.column_dimensions['H'].width = 30   # Generation Params
    
    # Save workbook
    wb.save(output_path)
    print(f"Excel file exported to: {output_path}")


# ============================================================================
# Main Workflow
# ============================================================================

async def main():
    """Main workflow: PDF -> Dataset -> Excel"""
    
    # ============================================================================
    # Interactive Input
    # ============================================================================
    console.print("\n" + "="*80)
    console.print("[bold cyan]Test Dataset Generation[/bold cyan]")
    console.print("="*80 + "\n")
    
    # Get PDF path
    console.print("[bold]请输入PDF文件路径:[/bold]")
    pdf_input = input("> ").strip()
    
    if not pdf_input:
        console.print("[red]✗ Error: PDF path cannot be empty[/red]")
        return
    
    # Remove quotes if present (from drag-and-drop in terminal)
    pdf_input = pdf_input.strip('"').strip("'")
    
    pdf_path = Path(pdf_input)
    if not pdf_path.exists():
        console.print(f"[red]✗ Error: PDF file not found: {pdf_path}[/red]")
        return
    
    # Get number of test cases
    console.print("\n[bold]请输入生成的测试用例数量 (默认: 50):[/bold]")
    num_cases_input = input("> ").strip()
    num_cases = int(num_cases_input) if num_cases_input else 50
    
    # Get max units
    console.print("\n[bold]请输入用于生成的最大单元数 (默认: 50):[/bold]")
    max_units_input = input("> ").strip()
    max_units = int(max_units_input) if max_units_input else 50
    
    # Generate output path (same directory as PDF, same name with .xlsx extension)
    output_path = pdf_path.parent / f"{pdf_path.stem}.xlsx"
    
    # ============================================================================
    # Setup
    # ============================================================================
    console.print("\n" + "="*80)
    console.print("[bold cyan]开始生成测试数据集[/bold cyan]")
    console.print("="*80 + "\n")
    
    # Check API key
    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        console.print("[red]✗ Error: OPENAI_API_KEY not found in environment[/red]")
        console.print("  Please set it in your .env file or environment")
        return
    
    llm_uri = "openai/gpt-4o-mini"
    
    console.print(Panel.fit(
        f"[bold]Configuration[/bold]\n\n"
        f"PDF File: {pdf_path}\n"
        f"LLM: {llm_uri}\n"
        f"Output: {output_path}\n"
        f"Num Cases: {num_cases}\n"
        f"Max Units: {max_units}",
        title="⚙️ Setup",
        border_style="blue"
    ))
    
    # ============================================================================
    # Stage 1: Read PDF
    # ============================================================================
    console.print("\n[bold yellow]Stage 1: Reading PDF[/bold yellow]")
    console.print("-" * 80)
    
    console.print("📖 Loading PDF with DoclingReader...")
    
    from docling.datamodel.pipeline_options import PdfPipelineOptions
    from docling.datamodel.accelerator_options import AcceleratorDevice, AcceleratorOptions
    
    # Use CPU for compatibility
    pdf_options = PdfPipelineOptions()
    pdf_options.accelerator_options = AcceleratorOptions(
        num_threads=8,
        device=AcceleratorDevice.CPU
    )
    
    reader = DoclingReader(pdf_pipeline_options=pdf_options)
    document = reader.read(str(pdf_path))
    
    console.print(f"[green]✓[/green] Document loaded")
    console.print(f"  - Content length: {len(document.content):,} characters")
    console.print(f"  - Pages: {len(document.pages)}")
    console.print(f"  - File: {document.metadata.file_name}")
    
    # ============================================================================
    # Stage 2: Split into Units
    # ============================================================================
    console.print("\n[bold yellow]Stage 2: Splitting Document[/bold yellow]")
    console.print("-" * 80)
    
    console.print("✂️  Splitting by markdown headers...")
    
    splitter = MarkdownHeaderSplitter()
    units = document.split(splitter)
    
    console.print(f"[green]✓[/green] Split into {len(units)} units")
    
    # Show sample units
    console.print("\n  Sample units:")
    for i, unit in enumerate(units[:3], 1):
        preview = unit.content[:100].replace("\n", " ")
        console.print(f"    {i}. [{unit.metadata.context_path}] {preview}...")
    
    if len(units) > 3:
        console.print(f"    ... and {len(units) - 3} more units")
    
    # ============================================================================
    # Stage 3: Enrich Units
    # ============================================================================
    console.print("\n[bold yellow]Stage 3: Enriching Units[/bold yellow]")
    console.print("-" * 80)
    
    console.print("🔍 Extracting summaries, keyphrases, and entities...")
    
    extractor = (
        SummaryExtractor(model_uri=llm_uri, api_key=api_key, max_sentences=2)
        | KeyphrasesExtractor(model_uri=llm_uri, api_key=api_key, max_num=5)
        | EntitiesExtractor(model_uri=llm_uri, api_key=api_key, max_num=5)
    )
    
    enriched_units = await extractor.transform(units, max_concurrency=3)
    
    console.print(f"[green]✓[/green] Enriched {len(enriched_units)} units")
    
    # Show enrichment sample
    if enriched_units:
        sample = enriched_units[0]
        console.print("\n  Sample enriched unit:")
        console.print(f"    Path: {sample.metadata.context_path}")
        if sample.summary:
            console.print(f"    Summary: {sample.summary[:100]}...")
        if sample.keyphrases:
            console.print(f"    Keyphrases: {', '.join(sample.keyphrases[:3])}")
        if sample.entities:
            console.print(f"    Entities: {', '.join(sample.entities[:3])}")
    
    # ============================================================================
    # Stage 4: Generate Personas
    # ============================================================================
    console.print("\n[bold yellow]Stage 4: Generating Personas[/bold yellow]")
    console.print("-" * 80)
    
    console.print("👥 Creating diverse user personas...")
    
    # Customize domain for your use case
    domain = "US residential real estate and home buying process"
    
    personas = await generate_personas(
        llm_uri=llm_uri,
        api_key=api_key,
        domain=domain,
        num_personas=3,
        persona_model=HomeBuyerPersona
    )
    
    console.print(f"[green]✓[/green] Generated {len(personas)} personas")
    for p in personas:
        console.print(f"  • [cyan]{p.name}[/cyan] - {p.expertise_level}")
    
    # ============================================================================
    # Stage 5: Generate Test Dataset
    # ============================================================================
    console.print("\n[bold yellow]Stage 5: Generating Test Dataset[/bold yellow]")
    console.print("-" * 80)
    
    console.print("💭 Generating single-hop questions...")
    
    # Limit units for generation
    max_units_to_use = min(max_units, len(enriched_units))
    units_for_generation = enriched_units[:max_units_to_use]
    
    console.print(f"  Using {len(units_for_generation)} units for generation")
    
    dataset = await generate_single_hop(
        llm_uri=llm_uri,
        api_key=api_key,
        units=units_for_generation,
        personas=personas,
        num_cases=num_cases,
        domain=domain
    )
    
    console.print(f"[green]✓[/green] Generated {len(dataset.cases)} test cases")
    
    # Show sample questions
    console.print("\n  Sample questions:")
    for i, case in enumerate(dataset.cases[:3], 1):
        console.print(f"    {i}. {case.question[:80]}...")
    
    # ============================================================================
    # Stage 6: Export to Excel
    # ============================================================================
    console.print("\n[bold yellow]Stage 6: Exporting to Excel[/bold yellow]")
    console.print("-" * 80)
    
    console.print("📊 Creating Excel file...")
    
    export_dataset_to_excel(dataset, output_path)
    
    console.print(f"[green]✓[/green] Excel file created")
    
    # ============================================================================
    # Summary
    # ============================================================================
    console.print("\n" + "="*80)
    console.print("[bold green]✨ Dataset Generation Completed![/bold green]")
    console.print("="*80 + "\n")
    
    console.print(Panel.fit(
        f"[bold]Summary[/bold]\n\n"
        f"✓ Document pages: {len(document.pages)}\n"
        f"✓ Units created: {len(units)}\n"
        f"✓ Units enriched: {len(enriched_units)}\n"
        f"✓ Personas generated: {len(personas)}\n"
        f"✓ Test cases: {len(dataset.cases)}\n\n"
        f"[cyan]Output file: {output_path}[/cyan]",
        title="📊 Results",
        border_style="green"
    ))
    
    console.print("\n[dim]Next steps:[/dim]")
    console.print(f"  Open the Excel file: {output_path}")
    console.print(f"  Review the generated test cases")
    console.print(f"  Use this dataset for RAG system evaluation")
    console.print()


if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        console.print("\n\n[yellow]Process interrupted by user[/yellow]")
    except Exception as e:
        console.print(f"\n\n[red]Error: {e}[/red]")
        import traceback
        console.print(traceback.format_exc())
